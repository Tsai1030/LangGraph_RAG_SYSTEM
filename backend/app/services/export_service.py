import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14)
_SUBTITLE_FONT = Font(italic=True, size=10, color="595959")
_NOTES_FONT = Font(size=9, color="595959")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_excel(form_data: dict, filename: str = "export") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = filename[:31]

    columns: list[str] = form_data.get("columns", [])
    rows: list[dict] = form_data.get("rows", [])
    title: str = form_data.get("title", filename)
    subtitle: str | None = form_data.get("subtitle")
    notes: str | None = form_data.get("notes")

    col_count = max(len(columns), 1)
    current_row = 1

    # Title row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=col_count)
    title_cell = ws.cell(row=current_row, column=1, value=title)
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _CENTER
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Subtitle row
    if subtitle:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=col_count)
        sub_cell = ws.cell(row=current_row, column=1, value=subtitle)
        sub_cell.font = _SUBTITLE_FONT
        sub_cell.alignment = _CENTER
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # Blank separator
    current_row += 1

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # Data rows
    for row_data in rows:
        for col_idx, col_name in enumerate(columns, start=1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.alignment = _LEFT
            cell.border = _THIN_BORDER
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # Notes
    if notes:
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=col_count)
        note_cell = ws.cell(row=current_row, column=1, value=f"備註：{notes}")
        note_cell.font = _NOTES_FONT
        note_cell.alignment = _LEFT

    # Auto column width (estimate from header + sample data)
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row_data in rows:
            val = str(row_data.get(col_name, ""))
            # Chinese chars are ~2x wide, approximate with *1.5
            length = sum(2 if ord(c) > 127 else 1 for c in val)
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_csv(form_data: dict) -> bytes:
    columns: list[str] = form_data.get("columns", [])
    rows: list[dict] = form_data.get("rows", [])

    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction="ignore", lineterminator="\r\n")
    writer.writeheader()
    writer.writerows(rows)

    # UTF-8 BOM so Excel opens Chinese correctly
    return "\ufeff".encode("utf-8") + buffer.getvalue().encode("utf-8")
